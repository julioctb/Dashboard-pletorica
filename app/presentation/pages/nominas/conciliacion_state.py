"""
Estado de la herramienta de conciliación nómina vs CONTPAQi.

Permite capturar manualmente los totales de CONTPAQi por empleado,
calcular diferencias y generar un semáforo visual de conciliación.
El Excel de exportación se sube a Supabase Storage y se descarga vía URL firmada.

Acceso: es_rrhh | es_contabilidad | es_admin_empresa
"""
import io
import logging
from datetime import date
from typing import Optional

import reflex as rx

from app.presentation.pages.nominas.base_state import NominaBaseState
from app.database import db_manager
from app.services.nomina_periodo_service import nomina_periodo_service

logger = logging.getLogger(__name__)

# Tolerancias para el semáforo de conciliación
_TOLERANCIA_VERDE = 1.0    # diferencia < $1 → verde
_TOLERANCIA_AMARILLO = 10.0  # diferencia < $10 → amarillo; >= $10 → rojo


def _calcular_semaforo(diff: float) -> str:
    """Verde <$1, Amarillo <$10, Rojo >=$10. Gris si no hay datos CONTPAQi."""
    if diff < 0:
        return 'gris'
    if diff < _TOLERANCIA_VERDE:
        return 'verde'
    if diff < _TOLERANCIA_AMARILLO:
        return 'amarillo'
    return 'rojo'


class NominaConciliacionState(NominaBaseState):
    """Estado de la herramienta de conciliación nómina ↔ CONTPAQi."""

    # Período activo
    periodo_id: str = ""
    periodos_disponibles: list[dict] = []

    # Lista de empleados: incluye sistema + contpaqi + semáforo
    empleados_conciliacion: list[dict] = []

    # Modal de captura CONTPAQi
    mostrar_modal_contpaqi: bool = False
    emp_editando_id: str = ""
    emp_editando_nombre: str = ""
    # Valores del sistema (solo lectura en el modal)
    modal_sistema_percepciones: float = 0.0
    modal_sistema_deducciones: float = 0.0
    modal_sistema_neto: float = 0.0
    # Inputs CONTPAQi
    form_contpaqi_percepciones: str = ""
    form_contpaqi_deducciones: str = ""
    form_contpaqi_neto: str = ""
    error_form: str = ""

    # URL de descarga del Excel
    url_excel: str = ""
    exportando: bool = False

    # =========================================================================
    # COMPUTED VARS — Semáforos
    # =========================================================================

    @rx.var
    def semaforo_verdes(self) -> int:
        return sum(1 for e in self.empleados_conciliacion if e.get('semaforo') == 'verde')

    @rx.var
    def semaforo_amarillos(self) -> int:
        return sum(
            1 for e in self.empleados_conciliacion if e.get('semaforo') == 'amarillo'
        )

    @rx.var
    def semaforo_rojos(self) -> int:
        return sum(1 for e in self.empleados_conciliacion if e.get('semaforo') == 'rojo')

    @rx.var
    def semaforo_grises(self) -> int:
        return sum(1 for e in self.empleados_conciliacion if e.get('semaforo') == 'gris')

    @rx.var
    def pct_cuadra(self) -> float:
        total = len(self.empleados_conciliacion)
        if total == 0:
            return 0.0
        return round(self.semaforo_verdes / total * 100, 1)

    @rx.var
    def tiene_empleados(self) -> bool:
        return len(self.empleados_conciliacion) > 0

    @rx.var
    def periodo_nombre_actual(self) -> str:
        for p in self.periodos_disponibles:
            if p.get('id') == self.periodo_id:
                return p.get('nombre', '')
        return ''

    # =========================================================================
    # MONTAJE
    # =========================================================================

    async def on_mount_conciliacion(self):
        if self.requiere_login and not self.esta_autenticado:
            yield rx.redirect("/login")
            return
        if not self.puede_acceder_nomina:
            yield rx.redirect(self.nomina_no_access_path)
            return
        if not self.id_empresa_actual:
            return
        await self._cargar_periodos_disponibles()

    # =========================================================================
    # HANDLERS — Selector de período
    # =========================================================================

    async def seleccionar_periodo_conciliacion(self, periodo_id: str):
        self.periodo_id = periodo_id
        self.empleados_conciliacion = []
        self.url_excel = ""
        if periodo_id:
            await self.cargar_conciliacion(int(periodo_id))

    # =========================================================================
    # HANDLER PRINCIPAL — Cargar conciliación
    # =========================================================================

    async def cargar_conciliacion(self, periodo_id: int):
        """Carga empleados del período con totales del sistema; CONTPAQi inicia vacío."""
        self.loading = True
        try:
            empleados = await nomina_periodo_service.obtener_empleados_periodo(periodo_id)
            empleados_ordenados = sorted(
                empleados,
                key=lambda item: (
                    item.get('nombre_empleado', ''),
                    item.get('clave_empleado', ''),
                ),
            )
            self.empleados_conciliacion = [
                {
                    **emp,
                    'id': str(emp['id']),
                    'cp_percepciones': 0.0,
                    'cp_deducciones': 0.0,
                    'cp_neto': 0.0,
                    'diff_neto': -1.0,   # -1 = sin datos CONTPAQi
                    'semaforo': 'gris',
                }
                for emp in empleados_ordenados
            ]
        except Exception as e:
            self.manejar_error(e, "cargar conciliación")
        finally:
            self.loading = False

    # =========================================================================
    # HANDLER — Abrir modal CONTPAQi
    # =========================================================================

    def abrir_modal_contpaqi(self, empleado: dict):
        self.emp_editando_id = str(empleado.get('id', ''))
        self.emp_editando_nombre = empleado.get('nombre_empleado', '')
        self.modal_sistema_percepciones = float(
            empleado.get('total_percepciones') or 0
        )
        self.modal_sistema_deducciones = float(
            empleado.get('total_deducciones') or 0
        )
        self.modal_sistema_neto = float(empleado.get('total_neto') or 0)
        # Pre-llenar con datos previos si existen
        self.form_contpaqi_percepciones = str(
            empleado.get('cp_percepciones') or ''
        ).replace('0.0', '')
        self.form_contpaqi_deducciones = str(
            empleado.get('cp_deducciones') or ''
        ).replace('0.0', '')
        self.form_contpaqi_neto = str(
            empleado.get('cp_neto') or ''
        ).replace('0.0', '')
        self.error_form = ""
        self.mostrar_modal_contpaqi = True

    def cerrar_modal_contpaqi(self):
        self.mostrar_modal_contpaqi = False
        self.emp_editando_id = ""
        self.error_form = ""

    def set_mostrar_modal_contpaqi(self, v: bool):
        self.mostrar_modal_contpaqi = v

    # Setters explícitos
    def set_form_contpaqi_percepciones(self, v: str):
        self.form_contpaqi_percepciones = v
        self.error_form = ""

    def set_form_contpaqi_deducciones(self, v: str):
        self.form_contpaqi_deducciones = v
        self.error_form = ""

    def set_form_contpaqi_neto(self, v: str):
        self.form_contpaqi_neto = v
        self.error_form = ""

    # =========================================================================
    # HANDLER — Guardar datos CONTPAQi
    # =========================================================================

    def capturar_contpaqi(self):
        """Almacena datos CONTPAQi en state y recalcula semáforo del empleado."""
        if not self.emp_editando_id:
            return
        try:
            cp = float(self.form_contpaqi_percepciones.replace(',', '') or 0)
            cd = float(self.form_contpaqi_deducciones.replace(',', '') or 0)
            cn = float(self.form_contpaqi_neto.replace(',', '') or 0)
        except ValueError:
            self.error_form = "Ingresa valores numéricos válidos (ej: 12500.50)"
            return

        if cn <= 0:
            self.error_form = "El neto CONTPAQi debe ser mayor a 0"
            return

        emp_id = self.emp_editando_id
        nueva_lista = []
        for emp in self.empleados_conciliacion:
            if str(emp.get('id', '')) == emp_id:
                sistema_neto = float(emp.get('total_neto') or 0)
                diff = abs(sistema_neto - cn)
                sem = _calcular_semaforo(diff)
                nueva_lista.append({
                    **emp,
                    'cp_percepciones': cp,
                    'cp_deducciones': cd,
                    'cp_neto': cn,
                    'diff_neto': round(diff, 2),
                    'semaforo': sem,
                })
            else:
                nueva_lista.append(emp)
        self.empleados_conciliacion = nueva_lista
        self.mostrar_modal_contpaqi = False

    # =========================================================================
    # EXPORTACIÓN EXCEL
    # =========================================================================

    async def exportar_excel(self):
        """Genera Excel con conciliación, sube a Storage y abre URL de descarga."""
        if not self.empleados_conciliacion:
            yield self.mostrar_mensaje("No hay datos para exportar.", "warning")
            return
        self.exportando = True
        try:
            contenido = self._generar_bytes_excel()
            nombre_periodo = self.periodo_nombre_actual.replace(' ', '_') or 'PERIODO'
            hoy = date.today().strftime('%Y%m%d')
            nombre_archivo = f"CONCILIACION_{nombre_periodo}_{hoy}.xlsx"
            storage_path = (
                f"nominas/{self.id_empresa_actual}/"
                f"{self.periodo_id}/{nombre_archivo}"
            )
            url = await self._subir_excel(storage_path, contenido)
            self.url_excel = url
            if url:
                yield rx.redirect(url, external=True)
                yield self.mostrar_mensaje("Excel generado correctamente.", "success")
            else:
                yield self.mostrar_mensaje("No se pudo subir el archivo.", "error")
        except Exception as e:
            self.manejar_error(e, "exportar Excel")
        finally:
            self.exportando = False

    def _generar_bytes_excel(self) -> bytes:
        """Construye el archivo Excel en memoria con openpyxl."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Hoja 1: Conciliación detallada ──────────────────────────────────
        ws = wb.active
        ws.title = "Conciliación"

        encabezados = [
            'Clave', 'Nombre',
            'Sistema — Percepciones', 'Sistema — Deducciones', 'Sistema — Neto',
            'CONTPAQi — Percepciones', 'CONTPAQi — Deducciones', 'CONTPAQi — Neto',
            'Diferencia Neto', 'Semáforo',
        ]
        header_fill = PatternFill(
            start_color='1E40AF', end_color='1E40AF', fill_type='solid'
        )
        header_font = Font(bold=True, color='FFFFFF')

        for col, h in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        semaforo_fills = {
            'verde':    PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
            'amarillo': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
            'rojo':     PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
        }

        for row_idx, emp in enumerate(self.empleados_conciliacion, 2):
            sem = emp.get('semaforo', 'gris')
            diff = float(emp.get('diff_neto') or -1)
            vals = [
                emp.get('clave_empleado', ''),
                emp.get('nombre_empleado', ''),
                float(emp.get('total_percepciones') or 0),
                float(emp.get('total_deducciones') or 0),
                float(emp.get('total_neto') or 0),
                float(emp.get('cp_percepciones') or 0),
                float(emp.get('cp_deducciones') or 0),
                float(emp.get('cp_neto') or 0),
                diff if diff >= 0 else 'Sin datos',
                sem.upper(),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if sem in semaforo_fills:
                    cell.fill = semaforo_fills[sem]

        # Ajuste de ancho de columnas
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value), default=10
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max_len + 4, 40
            )

        # ── Hoja 2: Resumen ──────────────────────────────────────────────────
        ws2 = wb.create_sheet("Resumen")
        nombre_p = self.periodo_nombre_actual or 'Período'
        ws2['A1'] = 'Período'
        ws2['B1'] = nombre_p
        ws2['A2'] = 'Total empleados'
        ws2['B2'] = len(self.empleados_conciliacion)
        ws2['A3'] = 'Total percepciones (sistema)'
        ws2['B3'] = sum(
            float(e.get('total_percepciones') or 0)
            for e in self.empleados_conciliacion
        )
        ws2['A4'] = 'Total deducciones (sistema)'
        ws2['B4'] = sum(
            float(e.get('total_deducciones') or 0)
            for e in self.empleados_conciliacion
        )
        ws2['A5'] = 'Total neto (sistema)'
        ws2['B5'] = sum(
            float(e.get('total_neto') or 0)
            for e in self.empleados_conciliacion
        )
        ws2['A7'] = 'Semáforo — Verde (< $1)'
        ws2['B7'] = self.semaforo_verdes
        ws2['A8'] = 'Semáforo — Amarillo (< $10)'
        ws2['B8'] = self.semaforo_amarillos
        ws2['A9'] = 'Semáforo — Rojo (>= $10)'
        ws2['B9'] = self.semaforo_rojos
        ws2['A10'] = '% que cuadra'
        ws2['B10'] = f"{self.pct_cuadra}%"

        for row in ws2.iter_rows(min_row=1, max_row=10, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def _subir_excel(self, storage_path: str, contenido: bytes) -> str:
        """Sube el Excel a Supabase Storage y retorna URL firmada (24h)."""
        supabase = db_manager.get_client()
        ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        try:
            supabase.storage.from_('archivos').upload(
                storage_path, contenido, {'content-type': ct, 'upsert': 'true'}
            )
        except Exception:
            try:
                supabase.storage.from_('archivos').update(
                    storage_path, contenido, {'content-type': ct}
                )
            except Exception as e2:
                logger.error(f"Error subiendo Excel a Storage: {e2}")
                return ''
        try:
            result = supabase.storage.from_('archivos').create_signed_url(
                storage_path, 86400
            )
            if isinstance(result, dict):
                return result.get('signedURL') or result.get('signedUrl', '')
        except Exception as e:
            logger.warning(f"No se pudo generar URL firmada para Excel: {e}")
        return ''

    # =========================================================================
    # CARGA DE PERÍODOS
    # =========================================================================

    async def _cargar_periodos_disponibles(self):
        self.loading = True
        try:
            supabase = db_manager.get_client()
            result = (
                supabase.table('periodos_nomina')
                .select('id, nombre, estatus, fecha_inicio')
                .eq('empresa_id', self.id_empresa_actual)
                .order('fecha_inicio', desc=True)
                .execute()
            )
            self.periodos_disponibles = [
                {**p, 'id': str(p['id'])} for p in (result.data or [])
            ]
            # Auto-seleccionar el más reciente si hay períodos cerrados/calculados
            if self.periodos_disponibles and not self.periodo_id:
                for p in self.periodos_disponibles:
                    if p.get('estatus') in ('CALCULADO', 'CERRADO'):
                        self.periodo_id = p['id']
                        await self.cargar_conciliacion(int(p['id']))
                        break
        except Exception as e:
            self.manejar_error(e, "cargar períodos para conciliación")
        finally:
            self.loading = False
