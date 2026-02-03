"""
Servicio de generacion de reportes para Alta Masiva.

Genera archivos Excel con los resultados del procesamiento,
con colores por tipo de resultado y hojas de resumen/detalle.
"""
import io
import logging
from datetime import datetime

from app.entities.alta_masiva import (
    ResultadoFila,
    ResultadoProcesamiento,
    ResultadoValidacion,
)

logger = logging.getLogger(__name__)


class ReporteAltaMasivaService:
    """Genera reportes Excel con resultados de alta masiva."""

    def generar_reporte_procesamiento(
        self,
        resultado: ResultadoProcesamiento,
        empresa_nombre: str = "",
    ) -> bytes:
        """
        Genera un reporte Excel con los resultados del procesamiento.

        Colores:
        - Verde: Empleado creado exitosamente
        - Amarillo: Reingreso exitoso
        - Rojo: Error en procesamiento

        Args:
            resultado: ResultadoProcesamiento con detalles
            empresa_nombre: Nombre de la empresa para el titulo

        Returns:
            Bytes del archivo .xlsx
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl no instalado")
            raise RuntimeError("Libreria openpyxl no instalada")

        wb = openpyxl.Workbook()

        # Estilos comunes
        title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        body_font = Font(name='Calibri', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Fills por resultado
        fill_creado = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        fill_reingreso = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        fill_error = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        font_creado = Font(name='Calibri', size=11, color='006100')
        font_reingreso = Font(name='Calibri', size=11, color='9C6500')
        font_error = Font(name='Calibri', size=11, color='9C0006')

        # --- Hoja Resumen ---
        ws_resumen = wb.active
        ws_resumen.title = 'Resumen'

        titulo = f'Reporte de Alta Masiva - {empresa_nombre}' if empresa_nombre else 'Reporte de Alta Masiva'
        ws_resumen.cell(row=1, column=1, value=titulo).font = title_font
        ws_resumen.cell(row=2, column=1, value=f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}').font = body_font

        ws_resumen.cell(row=4, column=1, value='Metrica').font = Font(name='Calibri', bold=True, size=11)
        ws_resumen.cell(row=4, column=2, value='Cantidad').font = Font(name='Calibri', bold=True, size=11)

        metricas = [
            ('Empleados creados', resultado.creados, fill_creado, font_creado),
            ('Reingresos procesados', resultado.reingresados, fill_reingreso, font_reingreso),
            ('Errores', resultado.errores, fill_error, font_error),
            ('Total procesados', resultado.total_procesados, None, Font(name='Calibri', bold=True, size=11)),
        ]

        for i, (metrica, valor, fill, font) in enumerate(metricas):
            row = 5 + i
            cell_m = ws_resumen.cell(row=row, column=1, value=metrica)
            cell_m.font = font
            cell_m.border = thin_border
            if fill:
                cell_m.fill = fill

            cell_v = ws_resumen.cell(row=row, column=2, value=valor)
            cell_v.font = font
            cell_v.border = thin_border
            cell_v.alignment = Alignment(horizontal='center')
            if fill:
                cell_v.fill = fill

        ws_resumen.column_dimensions['A'].width = 30
        ws_resumen.column_dimensions['B'].width = 15

        # --- Hoja Detalle ---
        ws_detalle = wb.create_sheet('Detalle')

        detalle_headers = ['Fila', 'CURP', 'Resultado', 'Clave', 'Mensaje']
        for col_idx, header in enumerate(detalle_headers, start=1):
            cell = ws_detalle.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        ws_detalle.column_dimensions['A'].width = 8
        ws_detalle.column_dimensions['B'].width = 22
        ws_detalle.column_dimensions['C'].width = 15
        ws_detalle.column_dimensions['D'].width = 14
        ws_detalle.column_dimensions['E'].width = 50

        for row_idx, detalle in enumerate(resultado.detalles, start=2):
            # Determinar estilo
            if detalle.resultado == ResultadoFila.VALIDO:
                fill = fill_creado
                font = font_creado
                resultado_texto = 'Creado'
            elif detalle.resultado == ResultadoFila.REINGRESO:
                fill = fill_reingreso
                font = font_reingreso
                resultado_texto = 'Reingreso'
            else:
                fill = fill_error
                font = font_error
                resultado_texto = 'Error'

            valores = [detalle.fila, detalle.curp, resultado_texto, detalle.clave, detalle.mensaje]
            for col_idx, valor in enumerate(valores, start=1):
                cell = ws_detalle.cell(row=row_idx, column=col_idx, value=valor)
                cell.font = font
                cell.fill = fill
                cell.border = thin_border

        # Congelar primera fila
        ws_detalle.freeze_panes = 'A2'

        # Guardar a bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generar_reporte_validacion(
        self,
        resultado: ResultadoValidacion,
        empresa_nombre: str = "",
    ) -> bytes:
        """
        Genera un reporte Excel con los resultados de la validacion (pre-procesamiento).

        Util para revisar errores antes de confirmar el procesamiento.

        Args:
            resultado: ResultadoValidacion con registros clasificados
            empresa_nombre: Nombre de la empresa

        Returns:
            Bytes del archivo .xlsx
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl no instalado")
            raise RuntimeError("Libreria openpyxl no instalada")

        wb = openpyxl.Workbook()

        # Estilos
        title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        body_font = Font(name='Calibri', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        fill_valido = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        fill_reingreso = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        fill_error = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # --- Hoja unica ---
        ws = wb.active
        ws.title = 'Validacion'

        titulo = f'Validacion de Alta Masiva - {empresa_nombre}' if empresa_nombre else 'Validacion de Alta Masiva'
        ws.cell(row=1, column=1, value=titulo).font = title_font
        ws.cell(row=2, column=1, value=f'Total filas: {resultado.total_filas}').font = body_font
        ws.cell(row=3, column=1,
                value=f'Validos: {resultado.total_validos} | Reingresos: {resultado.total_reingresos} | Errores: {resultado.total_errores}'
                ).font = body_font

        # Headers
        headers = ['Fila', 'CURP', 'Resultado', 'Mensaje']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 60

        # Combinar todos los registros y ordenar por fila
        todos = []
        for reg in resultado.validos:
            todos.append((reg, 'Valido', fill_valido))
        for reg in resultado.reingresos:
            todos.append((reg, 'Reingreso', fill_reingreso))
        for reg in resultado.errores:
            todos.append((reg, 'Error', fill_error))

        todos.sort(key=lambda x: x[0].fila)

        for row_idx, (reg, resultado_texto, fill) in enumerate(todos, start=6):
            mensaje = reg.mensaje or '; '.join(reg.errores)
            valores = [reg.fila, reg.curp, resultado_texto, mensaje]
            for col_idx, valor in enumerate(valores, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.font = body_font
                cell.fill = fill
                cell.border = thin_border

        ws.freeze_panes = 'A6'

        # Guardar a bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


# Singleton
reporte_alta_masiva_service = ReporteAltaMasivaService()
