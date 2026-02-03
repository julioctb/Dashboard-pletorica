"""
Servicio de generacion de plantillas para Alta Masiva.

Genera archivos Excel y CSV con el formato correcto para
la carga masiva de empleados.
"""
import io
import logging

logger = logging.getLogger(__name__)

# Definicion de columnas de la plantilla
COLUMNAS_PLANTILLA = [
    {'nombre': 'curp', 'titulo': 'CURP', 'ancho': 22, 'requerido': True,
     'instruccion': '18 caracteres. Obligatorio.'},
    {'nombre': 'nombre', 'titulo': 'Nombre', 'ancho': 20, 'requerido': True,
     'instruccion': 'Nombre(s) del empleado. Obligatorio.'},
    {'nombre': 'apellido_paterno', 'titulo': 'Apellido Paterno', 'ancho': 20, 'requerido': True,
     'instruccion': 'Primer apellido. Obligatorio.'},
    {'nombre': 'apellido_materno', 'titulo': 'Apellido Materno', 'ancho': 20, 'requerido': False,
     'instruccion': 'Segundo apellido. Opcional.'},
    {'nombre': 'rfc', 'titulo': 'RFC', 'ancho': 16, 'requerido': False,
     'instruccion': '13 caracteres (persona fisica). Opcional.'},
    {'nombre': 'nss', 'titulo': 'NSS', 'ancho': 14, 'requerido': False,
     'instruccion': '11 digitos (Seguro Social IMSS). Opcional.'},
    {'nombre': 'fecha_nacimiento', 'titulo': 'Fecha Nacimiento', 'ancho': 18, 'requerido': False,
     'instruccion': 'Formato DD/MM/AAAA. Opcional.'},
    {'nombre': 'genero', 'titulo': 'Genero', 'ancho': 12, 'requerido': False,
     'instruccion': 'Masculino o Femenino (M/F). Opcional.'},
    {'nombre': 'telefono', 'titulo': 'Telefono', 'ancho': 14, 'requerido': False,
     'instruccion': '10 digitos. Opcional.'},
    {'nombre': 'email', 'titulo': 'Email', 'ancho': 28, 'requerido': False,
     'instruccion': 'Correo electronico. Opcional.'},
    {'nombre': 'direccion', 'titulo': 'Direccion', 'ancho': 35, 'requerido': False,
     'instruccion': 'Domicilio completo. Opcional.'},
    {'nombre': 'contacto_emergencia', 'titulo': 'Contacto Emergencia', 'ancho': 30, 'requerido': False,
     'instruccion': 'Nombre y telefono. Opcional.'},
]


class PlantillaService:
    """Genera plantillas Excel y CSV para alta masiva."""

    def generar_excel(self) -> bytes:
        """
        Genera una plantilla Excel con:
        - Hoja 'Datos': Headers coloreados con formato
        - Hoja 'Instrucciones': Descripcion de cada columna

        Returns:
            Bytes del archivo .xlsx
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl no instalado")
            raise RuntimeError("Libreria openpyxl no instalada")

        wb = openpyxl.Workbook()

        # --- Hoja Datos ---
        ws_datos = wb.active
        ws_datos.title = 'Datos'

        # Estilos
        header_fill_req = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_fill_opt = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, col_def in enumerate(COLUMNAS_PLANTILLA, start=1):
            cell = ws_datos.cell(row=1, column=col_idx, value=col_def['titulo'])
            cell.font = header_font
            cell.fill = header_fill_req if col_def['requerido'] else header_fill_opt
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            ws_datos.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_def['ancho']

        # Fila de ejemplo
        ejemplo = {
            'curp': 'GARA850101HDFRZL09',
            'nombre': 'ALEJANDRO',
            'apellido_paterno': 'GARCIA',
            'apellido_materno': 'RAMIREZ',
            'rfc': 'GARA850101AB1',
            'nss': '12345678901',
            'fecha_nacimiento': '01/01/1985',
            'genero': 'Masculino',
            'telefono': '2221234567',
            'email': 'alejandro@ejemplo.com',
            'direccion': 'Av. Reforma 100, Col. Centro, Puebla',
            'contacto_emergencia': 'Maria Garcia 2229876543',
        }
        example_font = Font(name='Calibri', color='808080', italic=True, size=10)
        for col_idx, col_def in enumerate(COLUMNAS_PLANTILLA, start=1):
            cell = ws_datos.cell(row=2, column=col_idx, value=ejemplo.get(col_def['nombre'], ''))
            cell.font = example_font
            cell.border = thin_border

        # Congelar primera fila
        ws_datos.freeze_panes = 'A2'

        # --- Hoja Instrucciones ---
        ws_inst = wb.create_sheet('Instrucciones')

        inst_title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
        inst_header_font = Font(name='Calibri', bold=True, size=11)
        inst_body_font = Font(name='Calibri', size=11)

        ws_inst.cell(row=1, column=1, value='Instrucciones para Alta Masiva').font = inst_title_font
        ws_inst.cell(row=3, column=1, value='Columna').font = inst_header_font
        ws_inst.cell(row=3, column=2, value='Requerido').font = inst_header_font
        ws_inst.cell(row=3, column=3, value='Descripcion').font = inst_header_font

        ws_inst.column_dimensions['A'].width = 22
        ws_inst.column_dimensions['B'].width = 12
        ws_inst.column_dimensions['C'].width = 50

        for row_idx, col_def in enumerate(COLUMNAS_PLANTILLA, start=4):
            ws_inst.cell(row=row_idx, column=1, value=col_def['titulo']).font = inst_body_font
            req_text = 'Si' if col_def['requerido'] else 'No'
            ws_inst.cell(row=row_idx, column=2, value=req_text).font = inst_body_font
            ws_inst.cell(row=row_idx, column=3, value=col_def['instruccion']).font = inst_body_font

        notas_row = len(COLUMNAS_PLANTILLA) + 5
        ws_inst.cell(row=notas_row, column=1, value='Notas importantes:').font = inst_header_font
        notas = [
            'La fila 2 contiene un ejemplo que debe eliminar antes de subir.',
            'Maximo 500 registros por archivo.',
            'El CURP se usa para detectar empleados existentes (reingresos).',
            'Columnas azul oscuro son obligatorias, azul claro son opcionales.',
            'Los nombres se convierten a mayusculas automaticamente.',
        ]
        for i, nota in enumerate(notas):
            ws_inst.cell(row=notas_row + 1 + i, column=1, value=f'{i+1}. {nota}').font = inst_body_font

        # Guardar a bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generar_csv(self) -> bytes:
        """
        Genera una plantilla CSV con BOM UTF-8.

        Returns:
            Bytes del archivo CSV con BOM
        """
        headers = [col['titulo'] for col in COLUMNAS_PLANTILLA]
        linea = ','.join(headers)

        # BOM UTF-8 para que Excel abra correctamente
        bom = b'\xef\xbb\xbf'
        contenido = bom + (linea + '\n').encode('utf-8')

        return contenido


# Singleton
plantilla_service = PlantillaService()
