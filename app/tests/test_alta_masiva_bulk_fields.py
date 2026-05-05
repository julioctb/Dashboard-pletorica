"""Tests de campos extendidos para alta masiva de empleados."""

from datetime import date, datetime
from io import BytesIO

from app.domain.services.alta_masiva_parser import alta_masiva_parser
from app.domain.services.alta_masiva_service import AltaMasivaService
from app.domain.services.plantilla_service import plantilla_service
from app.presentation.components.shared.employee_bulk_upload_state_mixin import (
    EmployeeBulkUploadStateMixin,
)


def test_parser_reconoce_campos_extendidos_csv():
    contenido = (
        "CURP,Nombre,Apellido Paterno,Codigo Postal,Fecha Ingreso,"
        "Numero de Cuenta,Banco,Contacto Emergencia Nombre,"
        "Contacto Emergencia Telefono,Contacto Emergencia Parentesco\n"
        "GARA850101HDFRZL09,ALEJANDRO,GARCIA,72000,15/03/2024,"
        "1234567890,BBVA,MARIA GARCIA,2229876543,Padre/Madre\n"
    ).encode("utf-8")

    registros, errores = alta_masiva_parser.parsear(contenido, "empleados.csv")

    assert errores == []
    assert registros[0]["codigo_postal"] == "72000"
    assert registros[0]["fecha_ingreso"] == "15/03/2024"
    assert registros[0]["cuenta_bancaria"] == "1234567890"
    assert registros[0]["banco"] == "BBVA"
    assert registros[0]["contacto_emergencia_nombre"] == "MARIA GARCIA"
    assert registros[0]["contacto_emergencia_telefono"] == "2229876543"
    assert registros[0]["contacto_emergencia_parentesco"] == "Padre/Madre"


def test_parser_excel_convierte_fechas_a_formato_espanol_y_preserva_textos():
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "CURP",
            "Nombre",
            "Apellido Paterno",
            "Codigo Postal",
            "Fecha Ingreso",
            "NSS",
            "Cuenta Bancaria",
            "CLABE Interbancaria",
            "Contacto Emergencia Telefono",
        ]
    )
    sheet.append(
        [
            "GARA850101HDFRZL09",
            "ALEJANDRO",
            "GARCIA",
            "01234",
            datetime(2024, 3, 15),
            "01234567890",
            "001234567890",
            "002180000000000000",
            "0222987654",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)

    registros, errores = alta_masiva_parser.parsear(buffer.getvalue(), "empleados.xlsx")

    assert errores == []
    assert registros[0]["codigo_postal"] == "01234"
    assert registros[0]["fecha_ingreso"] == "15/03/2024"
    assert registros[0]["nss"] == "01234567890"
    assert registros[0]["cuenta_bancaria"] == "001234567890"
    assert registros[0]["clabe_interbancaria"] == "002180000000000000"
    assert registros[0]["contacto_emergencia_telefono"] == "0222987654"


def test_plantilla_excel_formatea_columnas_sensibles_como_texto():
    import openpyxl

    workbook = openpyxl.load_workbook(BytesIO(plantilla_service.generar_excel()))
    sheet = workbook["Datos"]
    headers = {cell.value: cell.column for cell in sheet[1]}

    for header in [
        "Codigo Postal",
        "Telefono",
        "NSS",
        "Cuenta Bancaria",
        "CLABE Interbancaria",
        "Contacto Emergencia Telefono",
    ]:
        assert sheet.cell(row=2, column=headers[header]).number_format == "@"
        assert sheet.cell(row=25, column=headers[header]).number_format == "@"

    assert (
        sheet.cell(row=2, column=headers["Fecha Ingreso"]).number_format == "dd/mm/yyyy"
    )


def test_crear_empleado_create_incluye_campos_extendidos_y_contacto_separado():
    service = AltaMasivaService()

    empleado = service._crear_empleado_create(
        {
            "curp": "GARA850101HDFRZL09",
            "nombre": "ALEJANDRO",
            "apellido_paterno": "GARCIA",
            "fecha_ingreso": "15/03/2024",
            "codigo_postal": "72000",
            "cuenta_bancaria": "1234567890",
            "banco": "bbva",
            "contacto_emergencia_nombre": "Maria Garcia",
            "contacto_emergencia_telefono": "2229876543",
            "contacto_emergencia_parentesco": "Padre/Madre",
        },
        empresa_id=7,
    )

    assert empleado.fecha_ingreso == date(2024, 3, 15)
    assert empleado.fecha_ingreso_vigente is None
    assert empleado.codigo_postal == "72000"
    assert empleado.cuenta_bancaria == "1234567890"
    assert empleado.banco == "BBVA"
    assert empleado.contacto_emergencia == "Maria Garcia / 2229876543 / Padre/Madre"


def test_crear_empleado_update_no_actualiza_fecha_ingreso_en_reingreso():
    service = AltaMasivaService()

    empleado_update = service._crear_empleado_update(
        {
            "fecha_ingreso": "15/03/2024",
            "codigo_postal": "72000",
            "cuenta_bancaria": "1234567890",
            "banco": "BBVA",
        }
    )

    data = empleado_update.model_dump(exclude_unset=True)
    assert "fecha_ingreso" not in data
    assert data["codigo_postal"] == "72000"
    assert data["cuenta_bancaria"] == "1234567890"
    assert data["banco"] == "BBVA"


def test_plantilla_csv_incluye_campos_extendidos():
    csv_data = plantilla_service.generar_csv().decode("utf-8-sig")

    assert "Codigo Postal" in csv_data
    assert "Fecha Ingreso" in csv_data
    assert "Cuenta Bancaria" in csv_data
    assert "CLABE Interbancaria" in csv_data
    assert "Contacto Emergencia Nombre" in csv_data
    assert "Contacto Emergencia Telefono" in csv_data
    assert "Contacto Emergencia Parentesco" in csv_data


def test_alta_masiva_paginar_helper_respeta_tamano_de_pagina():
    registros = [{"fila": fila} for fila in range(1, 26)]

    pagina_1 = EmployeeBulkUploadStateMixin._paginar_alta_masiva(
        registros,
        pagina=1,
        por_pagina=20,
    )
    pagina_2 = EmployeeBulkUploadStateMixin._paginar_alta_masiva(
        registros,
        pagina=2,
        por_pagina=20,
    )

    assert [item["fila"] for item in pagina_1] == list(range(1, 21))
    assert [item["fila"] for item in pagina_2] == list(range(21, 26))
    assert EmployeeBulkUploadStateMixin._calcular_total_paginas_alta_masiva(25, 20) == 2


def test_preview_reconstruye_registros_si_cache_esta_vacio():
    class BulkUploadStateForTest(EmployeeBulkUploadStateMixin):
        alta_masiva_validacion_validos = []
        alta_masiva_validacion_reingresos = []
        alta_masiva_validacion_errores = [
            EmployeeBulkUploadStateMixin._normalizar_registro_serializado(
                {
                    "fila": 2,
                    "resultado": "ERROR",
                    "errores": ["CURP invalida"],
                }
            )
        ]
    registros = BulkUploadStateForTest()._obtener_registros_preview_alta_masiva()

    assert len(registros) == 1
    assert registros[0]["fila"] == 2
    assert registros[0]["campo_error_display"] == "CURP"


def test_preview_materializa_filas_visibles_desde_errores():
    class BulkUploadStateForTest(EmployeeBulkUploadStateMixin):
        alta_masiva_validacion_validos = []
        alta_masiva_validacion_reingresos = []
        alta_masiva_validacion_errores = [
            EmployeeBulkUploadStateMixin._normalizar_registro_serializado(
                {
                    "fila": 2,
                    "resultado": "ERROR",
                    "errores": ["CLABE interbancaria invalida"],
                }
            )
        ]
        alta_masiva_preview_rows = []
        alta_masiva_preview_pagina = 1
        alta_masiva_por_pagina = 20

    state = BulkUploadStateForTest()
    state._actualizar_preview_paginado_alta_masiva()

    assert len(state.alta_masiva_preview_rows) == 1
    assert state.alta_masiva_preview_rows[0]["fila"] == 2
    assert state.alta_masiva_preview_rows[0]["campo_error_display"] == "CLABE Interbancaria"


def test_preview_materializa_solo_errores_y_excluye_procesables():
    class BulkUploadStateForTest(EmployeeBulkUploadStateMixin):
        alta_masiva_validacion_validos = [
            {"fila": 2, "resultado": "VALIDO", "curp": "GARA850101HDFRZL09"}
        ]
        alta_masiva_validacion_reingresos = [
            {"fila": 3, "resultado": "REINGRESO", "curp": "LOPE900101HDFRRS08"}
        ]
        alta_masiva_validacion_errores = [
            EmployeeBulkUploadStateMixin._normalizar_registro_serializado(
                {
                    "fila": 4,
                    "resultado": "ERROR",
                    "errores": ["NSS invalido"],
                }
            )
        ]
        alta_masiva_preview_rows = []
        alta_masiva_preview_pagina = 1
        alta_masiva_por_pagina = 20

    state = BulkUploadStateForTest()
    state._actualizar_preview_paginado_alta_masiva()

    assert [row["fila"] for row in state.alta_masiva_preview_rows] == [4]
    assert state.alta_masiva_resumen_paginacion_preview == "Mostrando 1-1 de 1 error(es)"


def test_alta_masiva_no_permite_procesar_si_hay_errores():
    class BulkUploadStateForTest(EmployeeBulkUploadStateMixin):
        alta_masiva_validacion_validos = [{"fila": 2, "resultado": "VALIDO"}]
        alta_masiva_validacion_reingresos = []
        alta_masiva_validacion_errores = [{"fila": 3, "resultado": "ERROR"}]

    assert BulkUploadStateForTest()._puede_procesar_alta_masiva() is False


def test_alta_masiva_permite_procesar_reingresos_sin_errores():
    class BulkUploadStateForTest(EmployeeBulkUploadStateMixin):
        alta_masiva_validacion_validos = []
        alta_masiva_validacion_reingresos = [{"fila": 2, "resultado": "REINGRESO"}]
        alta_masiva_validacion_errores = []

    assert BulkUploadStateForTest()._puede_procesar_alta_masiva() is True


def test_serializar_registros_agrega_campos_visibles_de_error():
    registro = EmployeeBulkUploadStateMixin._normalizar_registro_serializado(
        {
            "fila": 3,
            "curp": "",
            "resultado": "ERROR",
            "datos": {
                "curp": "LOPE900101HDFRRS08",
                "nombre": "JUAN",
                "apellido_paterno": "LOPEZ",
                "apellido_materno": "RUIZ",
            },
            "errores": ["CLABE interbancaria invalida"],
            "mensaje": "",
        }
    )

    assert registro["fila"] == 3
    assert registro["curp"] == "LOPE900101HDFRRS08"
    assert registro["nombre_completo"] == "JUAN LOPEZ RUIZ"
    assert registro["mensaje_display"] == "CLABE interbancaria invalida"
    assert registro["campo_error_display"] == "CLABE Interbancaria"


def test_serializar_registro_valido_no_muestra_campo_error():
    registro = EmployeeBulkUploadStateMixin._normalizar_registro_serializado(
        {
            "fila": 4,
            "curp": "GARA850101HDFRZL09",
            "resultado": "VALIDO",
            "datos": {
                "nombre": "ALEJANDRO",
                "apellido_paterno": "GARCIA",
            },
            "mensaje": "Registro valido para alta",
        }
    )

    assert registro["campo_error_display"] == "-"
    assert registro["mensaje_display"] == "Registro valido para alta"


def test_serializar_error_sin_detalle_mantiene_fallback_visible():
    registro = EmployeeBulkUploadStateMixin._normalizar_registro_serializado(
        {
            "fila": 5,
            "resultado": "ERROR",
            "datos": {},
            "errores": [],
            "mensaje": "",
        }
    )

    assert registro["mensaje_display"] == "Error de validacion sin detalle"
    assert registro["campo_error_display"] == "Validacion"


def test_validacion_alta_masiva_fecha_nacimiento_solo_valida_formato():
    service = AltaMasivaService()

    assert service._validar_fecha_nacimiento_alta_masiva("01/01/1900") == ""
    assert service._validar_fecha_nacimiento_alta_masiva("1900-01-01") == ""
    assert (
        "Fecha de nacimiento invalida"
        in service._validar_fecha_nacimiento_alta_masiva("01-01-1900")
    )


def test_validacion_alta_masiva_banco_permite_nombres_comerciales():
    service = AltaMasivaService()

    assert service._validar_banco_alta_masiva("BBVA MEXICO S.A.") == ""
    assert service._validar_banco_alta_masiva("BANCO & FIDEICOMISO-1") == ""
