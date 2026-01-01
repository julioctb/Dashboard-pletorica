"""
Calculadora de Costo Patronal - México
Versión Configurable - 2026

Permite ajustar todas las variables por empresa:
- Factor de integración personalizado
- Prima de riesgo específica
- Prestaciones superiores a la ley
- Días de pago por mes (30 o 30.4)

REFACTORIZACIÓN:
- Fase 1 (2025-12-31): Entidades movidas a app/entities/costo_patronal.py
- Fase 2 (2025-12-31): Calculadores especializados separados (IMSS, ISR, Provisiones)
- Orquestador: Este archivo coordina los calculadores especializados
"""

from typing import Optional
from app.core.fiscales import ConstantesFiscales, ISN_POR_ESTADO
from app.entities.costo_patronal import (
    ConfiguracionEmpresa,
    Trabajador,
    ResultadoCuotas
)
from app.core.calculations.calculadora_imss import CalculadoraIMSS
from app.core.calculations.calculadora_isr import CalculadoraISR
from app.core.calculations.calculadora_provisiones import CalculadoraProvisiones

# NOTA: Las clases ConfiguracionEmpresa, Trabajador y ResultadoCuotas
# fueron movidas a app/entities/costo_patronal.py en Fase 1 de refactorización


# =============================================================================
# CALCULADORA PRINCIPAL
# =============================================================================

class CalculadoraCostoPatronal:
    """
    Calculadora principal de costo patronal - Orquestador.

    Coordina los calculadores especializados:
    - CalculadoraIMSS: Cuotas IMSS patronales y obreras
    - CalculadoraISR: Impuesto Sobre la Renta
    - CalculadoraProvisiones: Aguinaldo, vacaciones, prima vacacional

    Responsabilidad: Ensamblar resultados de calculadores especializados
    en un objeto ResultadoCuotas completo.
    """

    def __init__(self, config: ConfiguracionEmpresa):
        """
        Inicializa calculadora con configuración de empresa.

        Args:
            config: Configuración personalizada de la empresa
        """
        self.config = config
        self.const = ConstantesFiscales
        self.resultados: list[ResultadoCuotas] = []

        # Inyectar calculadores especializados (Dependency Injection)
        self.calc_imss = CalculadoraIMSS(self.const)
        self.calc_isr = CalculadoraISR(self.const)
        self.calc_provisiones = CalculadoraProvisiones(self.const)
    
    def calcular(self, trabajador: Trabajador) -> ResultadoCuotas:
        """
        Calcula todas las cuotas para un trabajador.

        Orquesta los calculadores especializados y ensambla el resultado
        completo. Reducido de 129 líneas → 78 líneas mediante delegación.

        Args:
            trabajador: Datos del trabajador a calcular

        Returns:
            ResultadoCuotas con todos los campos calculados
        """
        dias = trabajador.dias_cotizados_mes

        # ═════════════════════════════════════════════════════════════════════
        # SALARIOS Y SBC
        # ═════════════════════════════════════════════════════════════════════
        factor_int = self.config.calcular_factor_integracion(trabajador.antiguedad_anos)
        sbc_diario = trabajador.salario_diario * factor_int
        sbc_diario = min(sbc_diario, self.const.TOPE_SBC_DIARIO)  # Aplicar tope
        salario_mensual = trabajador.salario_diario * dias

        # ═════════════════════════════════════════════════════════════════════
        # IMSS (delegar a calculadora especializada)
        # ═════════════════════════════════════════════════════════════════════
        imss_pat = self.calc_imss.calcular_patronal(
            sbc_diario,
            dias,
            self.config.prima_riesgo
        )

        es_sm = trabajador.es_salario_minimo(self.config.salario_minimo_aplicable)
        imss_obr, imss_absorbido = self.calc_imss.calcular_obrero(
            sbc_diario,
            dias,
            es_sm,
            self.config.aplicar_art_36_lss
        )

        # ═════════════════════════════════════════════════════════════════════
        # INFONAVIT (5% del SBC)
        # ═════════════════════════════════════════════════════════════════════
        infonavit = sbc_diario * self.const.INFONAVIT_PAT * dias

        # ═════════════════════════════════════════════════════════════════════
        # ISN (según estado configurado)
        # ═════════════════════════════════════════════════════════════════════
        isn = salario_mensual * self.config.tasa_isn

        # ═════════════════════════════════════════════════════════════════════
        # PROVISIONES (delegar a calculadora especializada)
        # ═════════════════════════════════════════════════════════════════════
        provisiones = self.calc_provisiones.calcular(
            trabajador.salario_diario,
            trabajador.antiguedad_anos,
            self.config.dias_aguinaldo,
            self.config.prima_vacacional,
            self.config.dias_vacaciones_adicionales
        )

        # ═════════════════════════════════════════════════════════════════════
        # ISR (delegar a calculadora especializada)
        # ═════════════════════════════════════════════════════════════════════
        isr = self.calc_isr.calcular(salario_mensual, es_sm)
        
        # ═════════════════════════════════════════════════════════════════════
        # RESULTADO (ensamblar desde dicts de calculadores especializados)
        # ═════════════════════════════════════════════════════════════════════
        return ResultadoCuotas(
            # Identificación
            trabajador=trabajador.nombre,
            empresa=self.config.nombre,

            # Salarios
            salario_diario=trabajador.salario_diario,
            salario_mensual=salario_mensual,
            factor_integracion=factor_int,
            sbc_diario=sbc_diario,
            sbc_mensual=sbc_diario * dias,
            dias_cotizados=dias,

            # IMSS Patronal (desempaquetar dict)
            imss_cuota_fija=imss_pat["cuota_fija"],
            imss_excedente_pat=imss_pat["excedente"],
            imss_prest_dinero_pat=imss_pat["prest_dinero"],
            imss_gastos_med_pens_pat=imss_pat["gastos_med"],
            imss_invalidez_vida_pat=imss_pat["invalidez_vida"],
            imss_guarderias=imss_pat["guarderias"],
            imss_retiro=imss_pat["retiro"],
            imss_cesantia_vejez_pat=imss_pat["cesantia_vejez"],
            imss_riesgo_trabajo=imss_pat["riesgo_trabajo"],

            # IMSS Obrero (desempaquetar dict)
            imss_excedente_obr=imss_obr["excedente"],
            imss_prest_dinero_obr=imss_obr["prest_dinero"],
            imss_gastos_med_pens_obr=imss_obr["gastos_med"],
            imss_invalidez_vida_obr=imss_obr["invalidez_vida"],
            imss_cesantia_vejez_obr=imss_obr["cesantia_vejez"],

            # Art. 36 LSS
            imss_obrero_absorbido=imss_absorbido,
            es_salario_minimo=es_sm,

            # Otros
            infonavit=infonavit,
            isn=isn,

            # Provisiones (desempaquetar dict)
            provision_aguinaldo=provisiones["aguinaldo"],
            provision_vacaciones=provisiones["vacaciones"],
            provision_prima_vac=provisiones["prima_vacacional"],

            # ISR (desempaquetar dict)
            isr_base_gravable=isr["base_gravable"],
            isr_antes_subsidio=isr["isr_antes_subsidio"],
            subsidio_empleo=isr["subsidio_empleo"],
            isr_a_retener=isr["isr_a_retener"],
        )
    
    def calcular_y_guardar(self, trabajador: Trabajador) -> ResultadoCuotas:
        resultado = self.calcular(trabajador)
        self.resultados.append(resultado)
        return resultado

    def calcular_desde_neto(self, salario_neto_deseado: float,
                            trabajador: Trabajador) -> tuple[ResultadoCuotas, int]:
        """
        Calcula el salario bruto necesario para alcanzar un neto deseado.
        Usa método iterativo (bisección) para encontrar el salario correcto.

        Retorna: (resultado, iteraciones)
        """
        # Límites de búsqueda
        salario_min = salario_neto_deseado / trabajador.dias_cotizados_mes  # Estimación baja
        salario_max = salario_neto_deseado * 2 / trabajador.dias_cotizados_mes  # Estimación alta

        max_iteraciones = 50
        tolerancia = 1.0  # $1 de tolerancia

        for i in range(max_iteraciones):
            # Probar con el punto medio
            salario_medio = (salario_min + salario_max) / 2

            # Crear trabajador temporal con este salario
            trabajador_temp = Trabajador(
                nombre=trabajador.nombre,
                salario_diario=salario_medio,
                antiguedad_anos=trabajador.antiguedad_anos,
                dias_cotizados_mes=trabajador.dias_cotizados_mes,
                zona_frontera=trabajador.zona_frontera,
            )

            # Calcular
            resultado = self.calcular(trabajador_temp)

            # Verificar si llegamos al neto deseado
            diferencia = resultado.salario_neto - salario_neto_deseado

            if abs(diferencia) <= tolerancia:
                return resultado, i + 1

            # Ajustar límites según el resultado
            if diferencia < 0:
                # El neto es menor, necesitamos salario bruto más alto
                salario_min = salario_medio
            else:
                # El neto es mayor, necesitamos salario bruto más bajo
                salario_max = salario_medio

        # Si no converge, retornar el mejor resultado
        return resultado, max_iteraciones

    def limpiar_resultados(self):
        """Limpia la lista de resultados guardados"""
        self.resultados.clear()

    def exportar_excel(self, nombre_archivo: str = "costo_patronal_2026.xlsx") -> str:
        """
        Exporta los resultados a Excel.
        Requiere la biblioteca openpyxl: pip install openpyxl

        Retorna: ruta del archivo creado
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "Para exportar a Excel necesitas instalar openpyxl:\n"
                "pip install openpyxl"
            )

        if not self.resultados:
            raise ValueError("No hay resultados para exportar")

        # Crear libro
        wb = openpyxl.Workbook()

        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        # Encabezados
        encabezados = [
            "Trabajador", "Salario Diario", "Salario Mensual", "SBC Diario",
            "IMSS Patronal", "IMSS Obrero", "INFONAVIT", "ISN",
            "Provisiones", "ISR", "Salario Neto", "Costo Total", "Factor Costo"
        ]

        for col, encabezado in enumerate(encabezados, 1):
            celda = ws_resumen.cell(1, col, encabezado)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            celda.alignment = Alignment(horizontal="center")

        # Datos
        for fila, r in enumerate(self.resultados, 2):
            ws_resumen.cell(fila, 1, r.trabajador)
            ws_resumen.cell(fila, 2, r.salario_diario)
            ws_resumen.cell(fila, 3, r.salario_mensual)
            ws_resumen.cell(fila, 4, r.sbc_diario)
            ws_resumen.cell(fila, 5, r.total_imss_patronal)
            ws_resumen.cell(fila, 6, r.total_imss_obrero)
            ws_resumen.cell(fila, 7, r.infonavit)
            ws_resumen.cell(fila, 8, r.isn)
            ws_resumen.cell(fila, 9, r.total_provisiones)
            ws_resumen.cell(fila, 10, r.isr_a_retener)
            ws_resumen.cell(fila, 11, r.salario_neto)
            ws_resumen.cell(fila, 12, r.costo_total)
            ws_resumen.cell(fila, 13, r.factor_costo)

        # Formatear columnas numéricas como moneda
        for fila in range(2, len(self.resultados) + 2):
            for col in range(2, 13):
                ws_resumen.cell(fila, col).number_format = '$#,##0.00'
            ws_resumen.cell(fila, 13).number_format = '0.0000'

        # Ajustar anchos
        for col in range(1, len(encabezados) + 1):
            ws_resumen.column_dimensions[get_column_letter(col)].width = 15

        # Hoja 2: Constantes
        ws_const = wb.create_sheet("Constantes 2026")
        ws_const['A1'] = "CONSTANTES FISCALES 2026"
        ws_const['A1'].font = Font(bold=True, size=14)

        fila = 3
        constantes_info = [
            ("SALARIOS MÍNIMOS", ""),
            ("  General", f"${self.const.SALARIO_MINIMO_GENERAL:.2f}/día"),
            ("  Frontera", f"${self.const.SALARIO_MINIMO_FRONTERA:.2f}/día"),
            ("", ""),
            ("UMA", ""),
            ("  Diario", f"${self.const.UMA_DIARIO:.2f}"),
            ("  Mensual", f"${self.const.UMA_MENSUAL:.2f}"),
            ("  Tope SBC", f"${self.const.TOPE_SBC_DIARIO:.2f}"),
            ("", ""),
            ("IMSS - TASAS", ""),
            ("  Cesantía y Vejez Patronal", f"{self.const.IMSS_CESANTIA_VEJEZ_PAT:.3%}"),
            ("  Cesantía y Vejez Obrero", f"{self.const.IMSS_CESANTIA_VEJEZ_OBR:.3%}"),
            ("  Retiro", f"{self.const.IMSS_RETIRO:.2%}"),
            ("  INFONAVIT", f"{self.const.INFONAVIT_PAT:.2%}"),
            ("", ""),
            ("SUBSIDIO AL EMPLEO", ""),
            ("  Mensual", f"${self.const.SUBSIDIO_EMPLEO_MENSUAL:.2f}"),
            ("  Límite", f"${self.const.LIMITE_SUBSIDIO_MENSUAL:.2f}"),
        ]

        for concepto, valor in constantes_info:
            ws_const.cell(fila, 1, concepto)
            ws_const.cell(fila, 2, valor)
            if concepto and not concepto.startswith("  "):
                ws_const.cell(fila, 1).font = Font(bold=True)
            fila += 1

        ws_const.column_dimensions['A'].width = 30
        ws_const.column_dimensions['B'].width = 20

        # Guardar archivo
        wb.save(nombre_archivo)
        return nombre_archivo


# =============================================================================
# FUNCIONES DE IMPRESIÓN
# =============================================================================

def imprimir_resultado(r: ResultadoCuotas):
    """Imprime resultado de forma legible"""
    
    print("\n" + "="*65)
    print(f"COSTO PATRONAL: {r.trabajador}")
    print(f"Empresa: {r.empresa}")
    print("="*65)
    
    if r.es_salario_minimo:
        print("\n   ⚠️  SALARIO MÍNIMO:")
        print("       • Exento de retención ISR (Art. 96 LISR)")
        print("       • Cuota obrera IMSS la paga el patrón (Art. 36 LSS)")
    
    print(f"\n📋 SALARIO")
    print(f"   Salario diario:            ${r.salario_diario:>12,.2f}")
    print(f"   Días cotizados:            {r.dias_cotizados:>12}")
    print(f"   Salario mensual:           ${r.salario_mensual:>12,.2f}")
    
    print(f"\n📊 INTEGRACIÓN")
    print(f"   Factor de integración:     {r.factor_integracion:>13.4f}")
    print(f"   SBC diario:                ${r.sbc_diario:>12,.2f}")
    print(f"   SBC mensual:               ${r.sbc_mensual:>12,.2f}")
    
    print(f"\n🏥 IMSS PATRONAL")
    print(f"   Cuota fija (E.M.):         ${r.imss_cuota_fija:>12,.2f}")
    print(f"   Excedente (E.M.):          ${r.imss_excedente_pat:>12,.2f}")
    print(f"   Prest. en dinero (E.M.):   ${r.imss_prest_dinero_pat:>12,.2f}")
    print(f"   Gastos méd. pensionados:   ${r.imss_gastos_med_pens_pat:>12,.2f}")
    print(f"   Invalidez y vida:          ${r.imss_invalidez_vida_pat:>12,.2f}")
    print(f"   Guarderías:                ${r.imss_guarderias:>12,.2f}")
    print(f"   Retiro:                    ${r.imss_retiro:>12,.2f}")
    print(f"   Cesantía y vejez (3.513%): ${r.imss_cesantia_vejez_pat:>12,.2f}")
    print(f"   Riesgo de trabajo:         ${r.imss_riesgo_trabajo:>12,.2f}")
    print(f"   ─────────────────────────────────────────────")
    print(f"   TOTAL IMSS PATRONAL:       ${r.total_imss_patronal:>12,.2f}")
    
    if r.imss_obrero_absorbido > 0:
        print(f"\n⚖️  ART. 36 LSS - CUOTA OBRERA (patrón paga)")
        print(f"   IMSS obrero absorbido:     ${r.imss_obrero_absorbido:>12,.2f}")
    
    print(f"\n🏠 INFONAVIT (5%)")
    print(f"   Aportación patronal:       ${r.infonavit:>12,.2f}")
    
    print(f"\n💰 ISN")
    print(f"   Impuesto sobre nómina:     ${r.isn:>12,.2f}")
    
    print(f"\n📅 PROVISIONES MENSUALES")
    print(f"   Aguinaldo:                 ${r.provision_aguinaldo:>12,.2f}")
    print(f"   Vacaciones:                ${r.provision_vacaciones:>12,.2f}")
    print(f"   Prima vacacional:          ${r.provision_prima_vac:>12,.2f}")
    print(f"   ─────────────────────────────────────────────")
    print(f"   TOTAL PROVISIONES:         ${r.total_provisiones:>12,.2f}")
    
    print(f"\n👷 DESCUENTOS AL TRABAJADOR")
    if r.es_salario_minimo:
        print(f"   IMSS Obrero:               $        0.00  (Art. 36 LSS)")
        print(f"   ISR:                       $        0.00  (Art. 96 LISR)")
    else:
        print(f"   IMSS Obrero:               ${r.total_imss_obrero:>12,.2f}")
        print(f"   ISR a retener:             ${r.isr_a_retener:>12,.2f}")
    print(f"   ─────────────────────────────────────────────")
    print(f"   TOTAL DESCUENTOS:          ${r.total_descuentos_trabajador:>12,.2f}")
    
    print(f"\n" + "="*65)
    print(f"💵 COSTO TOTAL PATRONAL:      ${r.costo_total:>12,.2f}")
    print(f"📈 FACTOR DE COSTO:           {r.factor_costo:>13.4f}")
    print(f"   (El trabajador cuesta {r.factor_costo:.2%} de su salario nominal)")
    print("="*65)
    print(f"💰 SALARIO NETO TRABAJADOR:   ${r.salario_neto:>12,.2f}")
    print("="*65)


# =============================================================================
# CONFIGURACIONES PREDEFINIDAS
# =============================================================================

def crear_configuracion_mantiser() -> ConfiguracionEmpresa:
    """Configuración de Mantiser basada en EMA real"""
    return ConfiguracionEmpresa(
        nombre="Mantiser",
        registro_patronal="Y46-55423-10-3",
        estado="puebla",
        prima_riesgo=0.0259840,  # 2.5984%
        factor_integracion_fijo=1.0493,  # Del EMA real
        dias_aguinaldo=15,
        prima_vacacional=0.25,
        dias_vacaciones_adicionales=0,
        dias_por_mes=30,
        zona_frontera=False,
        aplicar_art_36_lss=True,
    )


def crear_configuracion_pletorica() -> ConfiguracionEmpresa:
    """Configuración de Pletórica (ajustar según datos reales)"""
    return ConfiguracionEmpresa(
        nombre="Pletórica",
        registro_patronal="XXXXXX",  # Completar
        estado="puebla",
        prima_riesgo=0.005,  # Prima mínima clase I (ajustar según real)
        factor_integracion_fijo=None,  # Calcular automáticamente
        dias_aguinaldo=15,
        prima_vacacional=0.25,
        dias_vacaciones_adicionales=0,
        dias_por_mes=30,
        zona_frontera=False,
        aplicar_art_36_lss=True,
    )


# =============================================================================
# EJEMPLO DE USO
# =============================================================================

if __name__ == "__main__":
    
    # ─────────────────────────────────────────────────────────────────────────
    # CONFIGURAR EMPRESA
    # ─────────────────────────────────────────────────────────────────────────
    
    mantiser = crear_configuracion_mantiser()
    mantiser.mostrar_configuracion()
    
    # ─────────────────────────────────────────────────────────────────────────
    # CREAR CALCULADORA
    # ─────────────────────────────────────────────────────────────────────────
    
    calc = CalculadoraCostoPatronal(mantiser)
    
    # ─────────────────────────────────────────────────────────────────────────
    # CALCULAR TRABAJADOR SM
    # ─────────────────────────────────────────────────────────────────────────
    
    trabajador_sm = Trabajador(
        nombre="Trabajador SM 2026 (1 año)",
        salario_diario=ConstantesFiscales.SALARIO_MINIMO_GENERAL,
        antiguedad_anos=1,
        dias_cotizados_mes=30,
    )
    
    resultado = calc.calcular(trabajador_sm)
    imprimir_resultado(resultado)
    
    # ─────────────────────────────────────────────────────────────────────────
    # COMPARACIÓN CON CIFRAS DEL CONTADOR
    # ─────────────────────────────────────────────────────────────────────────
    
    print("\n" + "="*65)
    print("COMPARACIÓN CON CIFRAS DEL CONTADOR (2026)")
    print("="*65)
    
    # Cifras del contador (por día, usando 30.4 días)
    cont_imss = 53.68
    cont_rcv = 27.87
    cont_infonavit = 16.53
    cont_total = 98.08
    
    # Mis cifras (por día)
    dias = resultado.dias_cotizados
    mi_imss_dia = (resultado.total_imss_patronal - resultado.imss_retiro -
                   resultado.imss_cesantia_vejez_pat) / dias
    mi_rcv_dia = (resultado.imss_retiro + resultado.imss_cesantia_vejez_pat) / dias
    mi_infonavit_dia = resultado.infonavit / dias
    mi_obrero_dia = resultado.imss_obrero_absorbido / dias
    
    print(f"\n{'Concepto':<25} {'Contador':>12} {'Mi cálculo':>12}")
    print("─"*50)
    print(f"{'IMSS (sin RCV)':<25} ${cont_imss:>10,.2f} ${mi_imss_dia + mi_obrero_dia:>10,.2f}")
    print(f"{'RCV':<25} ${cont_rcv:>10,.2f} ${mi_rcv_dia:>10,.2f}")
    print(f"{'INFONAVIT':<25} ${cont_infonavit:>10,.2f} ${mi_infonavit_dia:>10,.2f}")
    print("─"*50)
    
    mi_total_dia = mi_imss_dia + mi_obrero_dia + mi_rcv_dia + mi_infonavit_dia
    print(f"{'TOTAL/día':<25} ${cont_total:>10,.2f} ${mi_total_dia:>10,.2f}")
    print(f"{'TOTAL/mes (×30.4)':<25} ${cont_total * 30.4:>10,.2f} ${mi_total_dia * 30.4:>10,.2f}")