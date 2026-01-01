"""
Exportador de resultados a Excel.

Responsabilidad: Generar archivos Excel con resultados de costo patronal,
incluyendo hoja de resumen y hoja de constantes fiscales.

Separado del código de cálculo para mantener Single Responsibility Principle.
El exportador es infraestructura, no lógica de negocio.

Fecha: 2025-12-31 (Fase 3 de refactorización)
"""

from typing import List
from app.entities.costo_patronal import ResultadoCuotas
from app.core.fiscales import ConstantesFiscales


class ExcelExporter:
    """
    Exporta resultados de costo patronal a archivos Excel.

    Genera dos hojas:
    1. Resumen: Tabla con resultados de todos los trabajadores
    2. Constantes 2026: Constantes fiscales usadas en los cálculos

    Requiere: openpyxl instalado (pip install openpyxl)
    """

    def __init__(self, constantes: type[ConstantesFiscales] = ConstantesFiscales):
        """
        Inicializa exportador con constantes fiscales.

        Args:
            constantes: Clase con constantes fiscales para hoja de referencia
        """
        self.const = constantes

    def exportar(
        self,
        resultados: List[ResultadoCuotas],
        nombre_archivo: str = "costo_patronal_2026.xlsx"
    ) -> str:
        """
        Exporta lista de resultados a archivo Excel.

        Args:
            resultados: Lista de ResultadoCuotas a exportar
            nombre_archivo: Nombre del archivo a crear (default: costo_patronal_2026.xlsx)

        Returns:
            Ruta del archivo creado

        Raises:
            ImportError: Si openpyxl no está instalado
            ValueError: Si la lista de resultados está vacía

        Ejemplo:
            >>> from app.core.exporters import ExcelExporter
            >>> exporter = ExcelExporter()
            >>> archivo = exporter.exportar(resultados, "nomina_enero_2026.xlsx")
            >>> print(f"Archivo creado: {archivo}")
        """
        # Validar dependencias
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "Para exportar a Excel necesitas instalar openpyxl:\n"
                "pip install openpyxl"
            )

        # Validar datos
        if not resultados:
            raise ValueError("No hay resultados para exportar")

        # ═════════════════════════════════════════════════════════════════════
        # CREAR LIBRO Y HOJA DE RESUMEN
        # ═════════════════════════════════════════════════════════════════════
        wb = openpyxl.Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        # Encabezados con estilo
        encabezados = [
            "Trabajador", "Salario Diario", "Salario Mensual", "SBC Diario",
            "IMSS Patronal", "IMSS Obrero", "INFONAVIT", "ISN",
            "Provisiones", "ISR", "Salario Neto", "Costo Total", "Factor Costo"
        ]

        for col, encabezado in enumerate(encabezados, 1):
            celda = ws_resumen.cell(1, col, encabezado)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill(
                start_color="366092",
                end_color="366092",
                fill_type="solid"
            )
            celda.alignment = Alignment(horizontal="center")

        # ═════════════════════════════════════════════════════════════════════
        # DATOS DE RESULTADOS
        # ═════════════════════════════════════════════════════════════════════
        for fila, r in enumerate(resultados, 2):
            ws_resumen.cell(fila, 1, r.trabajador)
            ws_resumen.cell(fila, 2, r.salario_diario)
            ws_resumen.cell(fila, 3, r.salario_mensual)
            ws_resumen.cell(fila, 4, r.sbc_diario)
            ws_resumen.cell(fila, 5, r.total_imss_patronal)
            ws_resumen.cell(fila, 6, r.total_imss_obrero)
            ws_resumen.cell(fila, 7, r.infonavit)
            ws_resumen.cell(fila, 8, r.isn)
            ws_resumen.cell(fila, 9, r.total_provisiones)
            ws_resumen.cell(fila, 10, r.isr_a_retener)
            ws_resumen.cell(fila, 11, r.salario_neto)
            ws_resumen.cell(fila, 12, r.costo_total)
            ws_resumen.cell(fila, 13, r.factor_costo)

        # ═════════════════════════════════════════════════════════════════════
        # FORMATEAR COLUMNAS NUMÉRICAS
        # ═════════════════════════════════════════════════════════════════════
        for fila in range(2, len(resultados) + 2):
            # Moneda para columnas 2-12
            for col in range(2, 13):
                ws_resumen.cell(fila, col).number_format = '$#,##0.00'
            # Factor de costo como decimal (4 decimales)
            ws_resumen.cell(fila, 13).number_format = '0.0000'

        # Ajustar anchos de columnas
        for col in range(1, len(encabezados) + 1):
            ws_resumen.column_dimensions[get_column_letter(col)].width = 15

        # ═════════════════════════════════════════════════════════════════════
        # HOJA 2: CONSTANTES FISCALES 2026
        # ═════════════════════════════════════════════════════════════════════
        ws_const = wb.create_sheet("Constantes 2026")
        ws_const['A1'] = "CONSTANTES FISCALES 2026"
        ws_const['A1'].font = Font(bold=True, size=14)

        fila = 3
        constantes_info = [
            ("SALARIOS MÍNIMOS", ""),
            ("  General", f"${self.const.SALARIO_MINIMO_GENERAL:.2f}/día"),
            ("  Frontera", f"${self.const.SALARIO_MINIMO_FRONTERA:.2f}/día"),
            ("", ""),
            ("UMA", ""),
            ("  Diario", f"${self.const.UMA_DIARIO:.2f}"),
            ("  Mensual", f"${self.const.UMA_MENSUAL:.2f}"),
            ("  Tope SBC", f"${self.const.TOPE_SBC_DIARIO:.2f}"),
            ("", ""),
            ("IMSS - TASAS", ""),
            ("  Cesantía y Vejez Patronal", f"{self.const.IMSS_CESANTIA_VEJEZ_PAT:.3%}"),
            ("  Cesantía y Vejez Obrero", f"{self.const.IMSS_CESANTIA_VEJEZ_OBR:.3%}"),
            ("  Retiro", f"{self.const.IMSS_RETIRO:.2%}"),
            ("  INFONAVIT", f"{self.const.INFONAVIT_PAT:.2%}"),
            ("", ""),
            ("SUBSIDIO AL EMPLEO", ""),
            ("  Mensual", f"${self.const.SUBSIDIO_EMPLEO_MENSUAL:.2f}"),
            ("  Límite", f"${self.const.LIMITE_SUBSIDIO_MENSUAL:.2f}"),
        ]

        for concepto, valor in constantes_info:
            ws_const.cell(fila, 1, concepto)
            ws_const.cell(fila, 2, valor)
            # Resaltar secciones principales
            if concepto and not concepto.startswith("  "):
                ws_const.cell(fila, 1).font = Font(bold=True)
            fila += 1

        # Ajustar anchos
        ws_const.column_dimensions['A'].width = 30
        ws_const.column_dimensions['B'].width = 20

        # ═════════════════════════════════════════════════════════════════════
        # GUARDAR ARCHIVO
        # ═════════════════════════════════════════════════════════════════════
        wb.save(nombre_archivo)
        return nombre_archivo
